"""Export service for generating XLSX files"""
import os
import tempfile
import gc
from typing import Tuple, Optional
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.cell import WriteOnlyCell
from services.firebase_service import FirebaseService
import logging

logger = logging.getLogger(__name__)

class ExportService:
    """Handles data export to XLSX format with memory optimization"""
    
    def __init__(self):
        """Initialize export service"""
        self.firebase_service = FirebaseService()
    
    def generate_xlsx(self, 
                     device_id: str, 
                     start_timestamp: int, 
                     end_timestamp: int,
                     include_manual: bool = True,
                     manual_only: bool = False) -> Tuple[Optional[str], Optional[str]]:
        """
        Generate XLSX file with automatic and/or manual sensor data using memory-optimized write-only mode
        
        Args:
            device_id: Device identifier
            start_timestamp: Start time in milliseconds
            end_timestamp: End time in milliseconds
            include_manual: Include manual data (default: True)
            manual_only: Export only manual data (default: False)
            
        Returns:
            Tuple of (file_path: str, error: str)
        """
        try:
            logger.info(f"Starting XLSX generation for device {device_id} (include_manual={include_manual}, manual_only={manual_only})")
            
            # Get automatic telemetry data (unless manual_only)
            telemetry_data = [] if manual_only else self.firebase_service.get_telemetry_data(
                device_id, 
                start_timestamp, 
                end_timestamp
            )
            
            # Get manual data (if requested)
            manual_data = [] if not include_manual else self.firebase_service.get_manual_data(
                device_id, 
                start_timestamp, 
                end_timestamp
            )
            
            if not telemetry_data and not manual_data:
                return None, "No data found for the specified period"
            
            logger.info(f"Retrieved {len(telemetry_data)} telemetry data points and {len(manual_data)} manual data points")
            
            # Create workbook in write-only mode for better memory efficiency
            wb = openpyxl.Workbook(write_only=True)
            
            # Automatic telemetry data grouped by sensor
            if telemetry_data:
                # Group data by sensor_id (in memory, but necessary for separate sheets)
                sensor_data = defaultdict(list)
                for entry in telemetry_data:
                    sensor_id = entry.get('sensor_id', 'unknown')
                    sensor_data[sensor_id].append(entry)
                
                # Create a tab for each sensor
                for sensor_id, entries in sensor_data.items():
                    logger.info(f"Processing automatic sensor {sensor_id} with {len(entries)} entries")
                    ws = wb.create_sheet(title=self._sanitize_sheet_name(sensor_id))
                    
                    # Get all unique value fields across all entries
                    value_fields = set()
                    for entry in entries:
                        value_fields.update(entry.get('values', {}).keys())
                    value_fields = sorted(list(value_fields))
                    
                    # Create header row with styling
                    headers = ['Timestamp', 'Date/Time', 'Metering Point', 'Sensor ID'] + value_fields
                    header_cells = []
                    for header_text in headers:
                        cell = WriteOnlyCell(ws, value=header_text)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        header_cells.append(cell)
                    ws.append(header_cells)
                    
                    # Add data rows (sorted by timestamp)
                    entries.sort(key=lambda x: x.get('timestamp', 0))
                    for entry in entries:
                        timestamp = entry.get('timestamp', 0)
                        dt = datetime.fromtimestamp(timestamp / 1000)
                        
                        row = [
                            timestamp,
                            dt.strftime('%Y-%m-%d %H:%M:%S'),
                            entry.get('metering_point', ''),
                            entry.get('sensor_id', '')
                        ]
                        
                        # Add value fields
                        values = entry.get('values', {})
                        for field in value_fields:
                            row.append(values.get(field, ''))
                        
                        ws.append(row)
                    
                    # Clear entries for this sensor to free memory
                    sensor_data[sensor_id] = None
                    gc.collect()
                    
                    logger.info(f"Completed automatic sensor {sensor_id}")
                
                # Clear sensor_data
                del sensor_data
                gc.collect()
            
            # Manual data
            if manual_data:
                ws_manual = wb.create_sheet(title="Manual")
                self._write_manual_sheet(ws_manual, manual_data)
                logger.info("Completed manual data sheet")
            
            # Clear original data lists to free memory
            del telemetry_data
            del manual_data
            gc.collect()
            
            # Save to temporary file
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, f'energiemonitor_{device_id}_{start_timestamp}.xlsx')
            
            logger.info(f"Saving XLSX to {file_path}")
            wb.save(file_path)
            
            # Clean up
            del wb
            gc.collect()
            
            logger.info("XLSX generation complete")
            return file_path, None
            
        except Exception as e:
            return None, f"Failed to generate XLSX: {str(e)}"
    
    def _write_combined_sheet(self, ws, telemetry_data: list, manual_data: list) -> None:
        """
        Write combined data (telemetry + manual) to Excel sheet.
        All data is sorted by timestamp for chronological view.
        
        Args:
            ws: Write-only worksheet
            telemetry_data: List of telemetry data points
            manual_data: List of manual data points
        """
        # Combine and sort by timestamp
        all_data = telemetry_data + manual_data
        all_data.sort(key=lambda x: x.get('timestamp', 0))
        
        # Headers
        headers = ['Timestamp', 'Date/Time', 'Type', 'Metering Point', 'Sensor ID', 'Data']
        header_cells = []
        for header_text in headers:
            cell = WriteOnlyCell(ws, value=header_text)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_cells.append(cell)
        ws.append(header_cells)
        
        # Data rows
        for point in all_data:
            timestamp = point.get('timestamp', 0)
            dt = datetime.fromtimestamp(timestamp / 1000)
            
            # Determine if manual or automatic
            is_manual = 'metadata' in point
            data_type = 'Manual' if is_manual else 'Automatic'
            
            # Format values as string
            values = point.get('values', {})
            values_str = ', '.join([f"{k}={v}" for k, v in values.items()])
            
            row = [
                timestamp,
                dt.strftime('%Y-%m-%d %H:%M:%S'),
                data_type,
                point.get('metering_point', ''),
                point.get('sensor_id', ''),
                values_str
            ]
            
            ws.append(row)
    
    def _write_manual_sheet(self, ws, manual_data: list) -> None:
        """
        Write manual data to Excel sheet with special formatting.
        
        Columns:
        - Timestamp
        - Date
        - Time
        - Metering Point
        - Energy Type
        - Description
        - Added Quantity
        - Leftover Quantity
        - Consumed Quantity
        - Unit
        - Purchase Date
        - Estimated Usage Date
        - Provisory
        
        Args:
            ws: Write-only worksheet
            manual_data: List of manual data points
        """
        # Headers
        headers = [
            'Timestamp', 'Date', 'Time', 'Metering Point', 
            'Energy Type', 'Description',
            'Added Quantity', 'Leftover Quantity', 'Consumed Quantity', 'Unit',
            'Purchase Date', 'Estimated Usage Date', 'Provisory'
        ]
        
        header_cells = []
        for header_text in headers:
            cell = WriteOnlyCell(ws, value=header_text)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_cells.append(cell)
        ws.append(header_cells)
        
        # Data rows
        for point in manual_data:
            timestamp = point.get('timestamp', 0)
            dt = datetime.fromtimestamp(timestamp / 1000)
            
            values = point.get('values', {})
            metadata = point.get('metadata', {})
            
            # Extract fields
            purchase_date_ts = metadata.get('purchase_date', timestamp)
            purchase_date = datetime.fromtimestamp(purchase_date_ts / 1000)
            
            usage_date_ts = metadata.get('estimated_usage_date', timestamp)
            usage_date = datetime.fromtimestamp(usage_date_ts / 1000)
            
            row = [
                timestamp,
                dt.strftime('%Y-%m-%d'),
                dt.strftime('%H:%M:%S'),
                point.get('metering_point', 'M0'),
                metadata.get('energy_type', 'unknown'),
                metadata.get('description', ''),
                values.get('added_quantity', 0),
                values.get('leftover_quantity', 0),
                values.get('consumed_quantity', 0),
                values.get('unit', ''),
                purchase_date.strftime('%Y-%m-%d %H:%M'),
                usage_date.strftime('%Y-%m-%d %H:%M'),
                'Ja' if metadata.get('usage_date_provisory', True) else 'Nein'
            ]
            
            ws.append(row)
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize sheet name to comply with Excel requirements
        Max 31 characters, no special characters
        """
        # Remove invalid characters
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Truncate to 31 characters
        return name[:31]
